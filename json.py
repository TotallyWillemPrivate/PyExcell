import openpyxl
import pprint
import re

geenTelnr = []

def excelToDict(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb[wb.sheetnames[0]]

    totalFile= []

    for row in range(2, sheet.max_row+1):
        newRow = {}
        for column in range(1, sheet.max_column+1):
            cellKey = sheet.cell(row=1, column=column).value
            cellValue = sheet.cell(row=row, column=column).value
            newRow[cellKey] = cellValue
        totalFile.append(newRow)

    return totalFile

def cleanFile(file):
    for record in file:
        #eerst postcode opmaken en controleren
        rePostcode = re.search(r'(\d{4})\s?([a-zA-Z]{2})', str(record['ADDR_ZIP']))
        if rePostcode == None:      
            if record['TELNR'] == None:
                record['ADDR_ZIP'] = 'FOUT'
        else:
            record['ADDR_ZIP'] = str(rePostcode[1]) + ' ' + str(rePostcode[2]).upper()
        
        #emailadressen allemaal in lowercase (als emailveld bestaat)
        mailAanwezig = 'EMAIL' in record.keys()
        if mailAanwezig == True:
            record['EMAIL'] = str(record['EMAIL']).lower()

        #telefoonnummers op maat maken (0 aan het begin weg halen) en verrijken (dit moet helemaal aan het eind?)
    return file

def checkPhone(file):
    for record in file:
        if record['TELNR'] != None:
            reTelnr = re.search(r'[0]*(\d*)\W?(\d*)', str(record['TELNR']))
            record['TELNR'] = int(reTelnr.group(1) + reTelnr.group(2))
            record['VERRIJKT_DOOR'] = 0
        #else:
            #api van CDDN
            #record['VERRIJKT_DOOR'] = 1
            #komt hier nog een nest met EDM
        else:
            #Op deze manier maak je 2 lijsten: een voor het nieuwe bestand en een voor het geentelnr bestand
            record['VERRIJKT_DOOR'] = 3
            geenTelnr.append(record)
            file.remove(record)
            
    return file

testfile = excelToDict('test.xlsx')
newFile = cleanFile(testfile)
newFile = checkPhone(newFile)
print('Nieuwe bestand')
pprint.pprint(newFile)
print('Geen telnr bestand')
pprint.pprint(geenTelnr)