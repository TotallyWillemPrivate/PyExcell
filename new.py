#Dit is een script waarin ik een voorzet schrijf voor een auto-verrijkingsdienst. 
#Ik probeer de excel-workflow op te zetten, voor de API-calls
import openpyxl
from openpyxl.utils import get_column_letter
import pprint
import re

#inladen van het excelbestand
wb = openpyxl.load_workbook('test.xlsx')
index = wb.sheetnames
sheet = wb[index[0]]

#aanmaken van een index voor de kolomkoppen
columnIndex = {}

#nieuwe kolom aanmaken voor verrijkingsnotering
verrijktcolumn = str(get_column_letter(sheet.max_column+1)) 
verrijktrow = str(1)
verrijktcell = verrijktcolumn+verrijktrow
sheet[verrijktcell] = 'VERRIJKT'

#vullen van de index met kolomkoppen
for column in range(1,sheet.max_column):
    columnName = sheet.cell(row=1, column=column).value
    columnIndex[columnName] = get_column_letter(column)

#Functie controlePostcode checkt of het een NL postcode is, als dat niet zo is en er is geen telefoonnummer aanwezig dan...
def controlePostcode():
    for row in range(2, sheet.max_row+1):
        cellZip = str(columnIndex['ADDR_ZIP']) + str(row)
        cellTelnr = str(columnIndex['TELNR']) + str(row)
        RegExPC = re.search(r'(\d{4})\s?([a-zA-Z]{2})', str(sheet[cellZip].value))
        if RegExPC == None:      
            if sheet[cellTelnr].value == None:
                sheet[cellZip] = 'FOUT'
        else:
            sheet[cellZip] = str(RegExPC[1]) + ' ' + str(RegExPC[2]).upper()

#Controleert of er een postcodeveld aanwezig is
if 'ADDR_ZIP' in columnIndex:
    controlePostcode()

#Per regel kijken in excel...
for row in range(2,sheet.max_row+1):
    cellTelnr = str(columnIndex['TELNR']) + str(row)
    
    #kijken of TELNR gevuld is
    if sheet[cellTelnr].value != None:
        #hier wordt een 0 ingevuld, omdat het adres is aangeleverd
        sheet[verrijktcolumn+str(row)] = '0'

    elif sheet[cellTelnr].value == None:
        #als dat niet zo is, dan doe je hier straks de API call
        if 'ADDR_ZIP' in columnIndex:
            cellZip = str(columnIndex['ADDR_ZIP']) + str(row)
            if sheet[cellZip].value != 'FOUT':
                sheet[cellTelnr] = 'checked'
                sheet[verrijktcolumn+str(row)] = '1'
            else:
                sheet[cellTelnr] = 'Postcode ook fout'
                sheet[verrijktcolumn+str(row)] = '3'
        else:
            sheet[cellTelnr] = 'geen postcodeveld?'

#opslaan van de nieuwe updated versie.
wb.save('updatedtext.xlsx')

""" 
Stappenplan:

API van CDDN aanspreken na de ELIF
De data van CDDN moet verwerkt worden
Daarna check op ontbrekende nummers
API van EDM aanspreken
De data van EDM met verwerkt worden

Dan komt het moeilijkste:
alle regels met een leeg telnr moeten worden gekopieerd naar een ander excel blad
(en het liefst verwijderd uit het oorspronkelijke)
Dat moet worden opgeslagen als 'geweigerd'
Kopie van oorspronkelijke moet worden opgeslagen als 'verrijkt' 

"""
